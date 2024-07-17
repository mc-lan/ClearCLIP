import openpyxl

def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet['A1'].value is None:
        sheet['A1'] = 'CLIP'
        sheet['B1'] = 'VIT'
        sheet['C1'] = 'MODEL'
        sheet['D1'] = 'Dataset'
        sheet['E1'] = 'aAcc'
        sheet['F1'] = 'mIoU'
        sheet['G1'] = 'mAcc'

    last_row = sheet.max_row

    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result['CLIP'])
        sheet.cell(row=last_row + index, column=2, value=result['VIT'])
        sheet.cell(row=last_row + index, column=3, value=result['MODEL'])
        sheet.cell(row=last_row + index, column=4, value=result['Dataset'])
        sheet.cell(row=last_row + index, column=5, value=result['aAcc'])
        sheet.cell(row=last_row + index, column=6, value=result['mIoU'])
        sheet.cell(row=last_row + index, column=7, value=result['mAcc'])

    workbook.save(file_path)

